"""Bounded, non-evaluating inspection of untrusted CSV and XLSX imports."""

import csv
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO, StringIO
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile, ZipFile

from defusedxml import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

type ImportScalar = str | int | float | bool | None


class ImportParseError(ValueError):
    """Safe parser failure whose reason contains no imported cell content."""

    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


@dataclass(frozen=True, slots=True)
class ImportColumnInspection:
    name: str
    sample_values: tuple[ImportScalar, ...]


@dataclass(frozen=True, slots=True)
class ImportSheetInspection:
    name: str
    row_count: int
    columns: tuple[ImportColumnInspection, ...]


@dataclass(frozen=True, slots=True)
class ImportInspection:
    sheets: tuple[ImportSheetInspection, ...]


@dataclass(frozen=True, slots=True)
class ImportSourceRow:
    sheet: str
    row_number: int
    values: dict[str, ImportScalar]


@dataclass(frozen=True, slots=True)
class ImportParserLimits:
    max_file_bytes: int = 25 * 1024 * 1024
    max_uncompressed_bytes: int = 200 * 1024 * 1024
    max_archive_entries: int = 2_000
    max_compression_ratio: int = 100
    max_sheets: int = 50
    max_rows_per_sheet: int = 100_000
    max_columns_per_sheet: int = 1_000
    max_cell_characters: int = 32_767
    sample_values_per_column: int = 5


class ImportParser:
    """Inspect import structure while enforcing deterministic resource limits."""

    def __init__(self, limits: ImportParserLimits | None = None) -> None:
        self._limits = limits or ImportParserLimits()

    def inspect(self, source: BinaryIO, *, filename: str) -> ImportInspection:
        suffix = Path(filename).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise ImportParseError("UNSUPPORTED_FILE_TYPE")
        payload = source.read(self._limits.max_file_bytes + 1)
        if len(payload) > self._limits.max_file_bytes:
            raise ImportParseError("FILE_TOO_LARGE")
        if not payload:
            raise ImportParseError("EMPTY_FILE")
        if suffix == ".csv":
            return self._inspect_csv(payload, filename=Path(filename).name)
        return self._inspect_xlsx(payload)

    def iter_rows(
        self, source: BinaryIO, *, filename: str, sheet_name: str | None = None
    ) -> Iterator[ImportSourceRow]:
        suffix = Path(filename).suffix.lower()
        payload = source.read(self._limits.max_file_bytes + 1)
        if len(payload) > self._limits.max_file_bytes:
            raise ImportParseError("FILE_TOO_LARGE")
        if not payload:
            raise ImportParseError("EMPTY_FILE")
        if suffix == ".csv":
            yield from self._iter_csv_rows(payload, filename=Path(filename).name)
            return
        if suffix != ".xlsx":
            raise ImportParseError("UNSUPPORTED_FILE_TYPE")
        yield from self._iter_xlsx_rows(payload, sheet_name=sheet_name)

    def _iter_csv_rows(self, payload: bytes, *, filename: str) -> Iterator[ImportSourceRow]:
        try:
            text = payload.decode("utf-8-sig")
            rows = csv.reader(StringIO(text, newline=""), dialect=self._csv_dialect(text))
            headers = self._validated_headers(next(rows))
            emitted = 0
            for row_number, row in enumerate(rows, start=2):
                if not any(value != "" for value in row):
                    continue
                emitted += 1
                if emitted > self._limits.max_rows_per_sheet:
                    raise ImportParseError("TOO_MANY_ROWS")
                yield ImportSourceRow(
                    filename,
                    row_number,
                    {
                        header: self._bounded_scalar(row[index] if index < len(row) else None)
                        for index, header in enumerate(headers)
                    },
                )
        except ImportParseError:
            raise
        except (csv.Error, StopIteration, UnicodeDecodeError) as error:
            raise ImportParseError("MALFORMED_CSV") from error

    def _iter_xlsx_rows(
        self, payload: bytes, *, sheet_name: str | None
    ) -> Iterator[ImportSourceRow]:
        self._validate_archive(payload)
        try:
            workbook = load_workbook(
                BytesIO(payload), read_only=True, data_only=True, keep_links=False
            )
            worksheets = (
                [workbook[sheet_name]]
                if sheet_name is not None and sheet_name in workbook.sheetnames
                else list(workbook.worksheets)
                if sheet_name is None
                else []
            )
            if not worksheets:
                raise ImportParseError("SHEET_NOT_FOUND")
            for worksheet in worksheets:
                rows = worksheet.iter_rows(values_only=True)
                headers = self._validated_headers(next(rows))
                emitted = 0
                for row_number, row in enumerate(rows, start=2):
                    if not any(value is not None and value != "" for value in row):
                        continue
                    emitted += 1
                    if emitted > self._limits.max_rows_per_sheet:
                        raise ImportParseError("TOO_MANY_ROWS")
                    yield ImportSourceRow(
                        worksheet.title,
                        row_number,
                        {
                            header: self._bounded_scalar(row[index] if index < len(row) else None)
                            for index, header in enumerate(headers)
                        },
                    )
        except ImportParseError:
            raise
        except (
            BadZipFile,
            InvalidFileException,
            DefusedXmlException,
            OSError,
            ValueError,
            StopIteration,
        ) as error:
            raise ImportParseError("MALFORMED_XLSX") from error
        finally:
            if "workbook" in locals():
                workbook.close()

    def _validated_headers(self, values: tuple[object, ...] | list[str]) -> tuple[str, ...]:
        if len(values) > self._limits.max_columns_per_sheet:
            raise ImportParseError("TOO_MANY_COLUMNS")
        headers = tuple(self._header(value) for value in values)
        if not headers or any(not value for value in headers):
            raise ImportParseError("INVALID_HEADER")
        if len(set(headers)) != len(headers):
            raise ImportParseError("DUPLICATE_HEADER")
        return headers

    def _header(self, value: object) -> str:
        normalized = self._bounded_scalar(value)
        return "" if normalized is None else str(normalized).strip()

    def _bounded_scalar(self, value: object) -> ImportScalar:
        if value is None or isinstance(value, (bool, int, float)):
            return value
        if isinstance(value, (date, datetime, time)):
            return value.isoformat()
        text = str(value)
        if len(text) > self._limits.max_cell_characters:
            raise ImportParseError("CELL_TOO_LARGE")
        return text

    def _inspect_csv(self, payload: bytes, *, filename: str) -> ImportInspection:
        try:
            text = payload.decode("utf-8-sig")
        except UnicodeDecodeError as error:
            raise ImportParseError("CSV_INVALID_ENCODING") from error
        try:
            dialect = self._csv_dialect(text)
            rows = csv.reader(StringIO(text, newline=""), dialect=dialect)
            header = next(rows)
            collector = _SheetCollector(filename, header, self._limits)
            for row in rows:
                collector.add_row(row)
            return ImportInspection(sheets=(collector.result(),))
        except ImportParseError:
            raise
        except (csv.Error, StopIteration) as error:
            raise ImportParseError("MALFORMED_CSV") from error

    @staticmethod
    def _csv_dialect(text: str) -> type[csv.Dialect] | csv.Dialect:
        sample = text[:16_384]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            return csv.excel

    def _inspect_xlsx(self, payload: bytes) -> ImportInspection:
        self._validate_archive(payload)
        try:
            workbook = load_workbook(
                BytesIO(payload), read_only=True, data_only=True, keep_links=False
            )
            if len(workbook.sheetnames) > self._limits.max_sheets:
                raise ImportParseError("TOO_MANY_SHEETS")
            sheets: list[ImportSheetInspection] = []
            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    raise ImportParseError("EMPTY_SHEET") from None
                collector = _SheetCollector(worksheet.title, header, self._limits)
                for row in rows:
                    collector.add_row(row)
                sheets.append(collector.result())
            if not sheets:
                raise ImportParseError("EMPTY_WORKBOOK")
            return ImportInspection(sheets=tuple(sheets))
        except ImportParseError:
            raise
        except (
            BadZipFile,
            InvalidFileException,
            DefusedXmlException,
            OSError,
            ValueError,
        ) as error:
            raise ImportParseError("MALFORMED_XLSX") from error
        finally:
            if "workbook" in locals():
                workbook.close()

    def _validate_archive(self, payload: bytes) -> None:
        try:
            with ZipFile(BytesIO(payload)) as archive:
                entries = archive.infolist()
                if len(entries) > self._limits.max_archive_entries:
                    raise ImportParseError("XLSX_ARCHIVE_TOO_LARGE")
                total_size = 0
                for entry in entries:
                    total_size += entry.file_size
                    if total_size > self._limits.max_uncompressed_bytes:
                        raise ImportParseError("XLSX_ARCHIVE_TOO_LARGE")
                    compressed_size = max(entry.compress_size, 1)
                    if entry.file_size / compressed_size > self._limits.max_compression_ratio:
                        raise ImportParseError("XLSX_SUSPICIOUS_COMPRESSION")
        except BadZipFile as error:
            raise ImportParseError("MALFORMED_XLSX") from error


class _SheetCollector:
    def __init__(
        self,
        name: str,
        header: tuple[object, ...] | list[str],
        limits: ImportParserLimits,
    ) -> None:
        if len(header) > limits.max_columns_per_sheet:
            raise ImportParseError("TOO_MANY_COLUMNS")
        self._limits = limits
        self._headers = tuple(self._header(value) for value in header)
        if not self._headers or any(not value for value in self._headers):
            raise ImportParseError("INVALID_HEADER")
        if len(set(self._headers)) != len(self._headers):
            raise ImportParseError("DUPLICATE_HEADER")
        self._name = name
        self._row_count = 0
        self._samples: list[list[ImportScalar]] = [[] for _ in self._headers]

    def add_row(self, values: tuple[object, ...] | list[str]) -> None:
        if not any(value is not None and value != "" for value in values):
            return
        self._row_count += 1
        if self._row_count > self._limits.max_rows_per_sheet:
            raise ImportParseError("TOO_MANY_ROWS")
        for index, value in enumerate(values[: len(self._headers)]):
            sample = self._scalar(value)
            if sample is None or sample == "":
                continue
            if len(self._samples[index]) < self._limits.sample_values_per_column:
                self._samples[index].append(sample)

    def result(self) -> ImportSheetInspection:
        return ImportSheetInspection(
            name=self._name,
            row_count=self._row_count,
            columns=tuple(
                ImportColumnInspection(name=header, sample_values=tuple(self._samples[index]))
                for index, header in enumerate(self._headers)
            ),
        )

    def _header(self, value: object) -> str:
        normalized = self._scalar(value)
        if normalized is None:
            return ""
        return str(normalized).strip()

    def _scalar(self, value: object) -> ImportScalar:
        if value is None or isinstance(value, (bool, int, float)):
            return value
        if isinstance(value, (date, datetime, time)):
            return value.isoformat()
        text = str(value)
        if len(text) > self._limits.max_cell_characters:
            raise ImportParseError("CELL_TOO_LARGE")
        return text
